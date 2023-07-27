import pandas as pd
import datetime
from openpyxl import load_workbook
from io import BytesIO
import openpyxl
import pytz

# Compute average of pulled CID confidence scores
def compute_average_score(docs):
    total = 0
    for doc in docs.values():
        total += doc.score
        avgscore = total / len(docs) ## convert total score to avg
    
    avgscore *= 100 ## convert from decimal to percentage
    return avgscore

# Formate SME names for display
def parse_sme_name(sme):
    if sme == "N/A":
        return "<insert name here>"
    
    # Split name, remove whitespace
    name_list = sme.replace('/', ',').split(',')
    name_list = [name.strip() for name in name_list]
    l = len(name_list)

    # Reorder
    if l > 2:  # Handle multiple names case
        firstnames = [name_list[i] for i in range(1, len(name_list)-1)]
        firstname = ' '.join(firstnames)
        middlename = name_list[-1].replace('(', '\"').replace(')', '\"').strip()
        lastname = name_list[0]
        fullname = firstname + ' ' + middlename + ' ' + lastname
    elif l == 2:  # Handle Firstname Lastname case
        firstname = name_list[1].replace('(', '\"').replace(')', '\"').strip()
        lastname = name_list[0]
        fullname = firstname + ' ' + lastname
    elif l == 1:
        sme = "STRANGE SME NAME. INSPECT ORIGINAL DOCUMENT"
    else:
        fullname = name_list[1] + ' ' + name_list[0]
    return fullname

# Remove duplicate docs? --> (Who ever made the function give context)
def remove_duplicates(original, arr1, arr2=[], arr3=[]):
    index_dict = {}

    # Find indexes of duplicates in original array
    for i, x in enumerate(original):
        if x not in index_dict:
            index_dict[x] = [i]
        else:
            index_dict[x].append(i)

    # Remove corresponding items from other arrays
    for indexes in index_dict.values():
        if len(indexes) > 1:
            for i in indexes[1:]:
                if i < len(arr1):
                    del arr1[i]
                if i < len(arr2):
                    del arr2[i]
                if i < len(arr3):
                    del arr3[i]
    unique = list(set(original))
    original = [x for x in unique]

    return original, arr1, arr2, arr3

# Read questions from file and gather row data
def read_questions(file):

    # Read CSV or Excel file into a Pandas DataFrame with no headers
    if file.type == 'text/csv':
        df = pd.read_csv(file, header=None)
    elif file.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        df = pd.read_excel(file, header=None)
    else:
        return [], 2
    
    # Extract non-empty cells as questions and gather row data
    questions = []
    rows = []
    for _, row in df.iterrows():
        question = ""
        for cell in row:
            if pd.notna(cell):
                question += (str(cell).strip() + " ")
        if question != "":
            questions.append(question)
            rows.append(row)
    if questions==[]: ## If not question
        return [], 1, []
    return questions, 0, rows

# Format excel sheet considering original rows questions where in
def to_excel(df, rows):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    workbook = writer.book
    worksheet = workbook.add_worksheet("Sheet1")
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None, format1)
    n = 0
    for row in rows:
        for column in range(5):
            worksheet.write(row.name, column, str(df.iloc[n,column]))
        n += 1
    writer.close()
    processed_data = output.getvalue()
    return processed_data

# Formats excel sheet to not consider original rows --> NOT IN USE
def to_excel_no_format(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=None, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None, format1)
    writer.close()
    processed_data = output.getvalue()
    return processed_data

# Compare dates
def getMostRecentDate(x, y):
    # Convert strings to datetime objects
    date_x = datetime.strptime(x, "%Y-%m-%d %H:%M:%S %z")
    date_y = datetime.strptime(y, "%Y-%m-%d %H:%M:%S %z")

    # Compare dates and print the more recent one
    if date_x > date_y:
        return date_x
    else:
        return date_y
    
# Clean confidences scores for display
def clean_confidences(confidences):
    for i in range(len(confidences)):
        x = confidences[i].find("** ")
        confidences[i] = confidences[i][x+3:]
    return confidences
                    
#def read_questions_v2(file):
#    wb = openpyxl.load_workbook(file)
#    ws = wb.worksheets[0]
#
#    data = [[cell.value for cell in row] for row in ws[range_str]]
#
#    validations = ws.data_validations.dataValidation



# def get_email_text(query, best_sme, email_header, email_content):
def get_email_text(query, best_sme):

    print("Drafting email...")
    # Get response from OpenAI
    # prompt = f"Please write a brief and professional business email to someone named {best_sme} asking {query}. Include only the text of the email in your response, not any sort of email address, and should be formatted nicely. The email should start with Subject: __ \n\nand end with the exact string \n\n'[Your Name]'."
    # response = openai.Completion.create(
    #     engine='immerse-3-5',
    #     prompt=prompt,
    #     temperature=0.3,
    #     max_tokens=400,
    #     frequency_penalty=0.0,
    #     presence_penalty=0,
    # )
    # Substitute sections of email text and write
    # email_response = response.choices[0].text
    # subject_index = email_response.find("Subject:")
    # name_index = email_response.find("[Your Name]")
    # email_response = email_response[subject_index:name_index+len("[Your Name]")].strip()
    # email_content.write(email_response)

    email_text = f"""Subject: [subject of question]

    Good [morning/afternoon] {best_sme},

    We are working on a Security request for one of our customers. Their Technical Questionnaire has a question(s) for which you are listed as the subject matter expert. Please provide a response to the following question, after which we will add it to our database of standard enterprise content.

    

    {query}

    

    Our time is limited. Please provide a response by no later than [time], [Day], [mm/dd].

    If you require more time, or if the question(s) need to be redirected please respond upon receipt of this email so we have adequate time to redirect to the appropriate person(s).

    If you have questions or require additional information about the customer, products and/or services please reach out.

    Kind Regards,

    

    [Your Name]"""
    return email_text